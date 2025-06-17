# source: 1
import torch
import torch.nn as nn
import numpy as np
import joblib
import matplotlib.pyplot as plt
from openpyxl import Workbook
# from mpl_toolkits.mplot3d import Axes3D # 取消注释以启用 3D 绘图

# --- 1. 与训练时完全相同的模型定义 ---
# (从 ML_Training_1.txt 或您最终使用的训练脚本中复制)
# source: 16, 17, 18
class MLPRegression(nn.Module):
    """一个简单的多层感知机用于回归"""
    def __init__(self, input_dim, output_dim, h1, h2, h3):
        super(MLPRegression, self).__init__()
        # !!! 确保这里的层结构和激活函数与训练时完全一致 !!!
        self.network = nn.Sequential(
            nn.Linear(input_dim, h1),
            nn.Tanh(), # 或者 nn.ReLU()
            nn.Linear(h1, h2),
            nn.Tanh(), # 或者 nn.ReLU()
            nn.Linear(h2, h3),
            nn.Tanh(), # 或者 nn.ReLU()
            nn.Linear(h3, output_dim)
        )

    def forward(self, x):
        return self.network(x)

# --- 2. 定义与训练时相同的参数 ---
# source: 2
INPUT_DIM = 1
NODE_COUNT = 63 # 节点数量
OUTPUT_DIM = NODE_COUNT * 3 # 189

# !!! 重要：这里的隐藏层大小必须与训练并保存 best_mlp_model.pth 时使用的完全一致 !!!
# source: 2, 3
HIDDEN_LAYER_1 = 128 # 例如 (根据您训练时的设置为准)
HIDDEN_LAYER_2 = 256 # 例如 (根据您训练时的设置为准)
HIDDEN_LAYER_3 = 128 # 例如 (根据您训练时的设置为准)

# --- 3. 定义文件路径 ---
MODEL_PATH = 'best_mlp_model.pth'
X_SCALER_PATH = 'x_scaler.joblib'
Y_SCALER_PATH = 'y_scaler.joblib'
# !!! 新增：指定包含 63 个节点初始坐标 (X0, Y0, Z0) 的文件路径 !!!
# 文件格式：每行 X Y Z，用空格分隔，共 63 行，顺序必须与模型输出对应
INITIAL_COORDS_PATH = 'initial_coordinates.txt'

# --- 4. 加载模型和 Scaler 的函数 (与之前相同) ---
def load_prediction_components(model_path, x_scaler_path, y_scaler_path, input_dim, output_dim, h1, h2, h3):
    """加载模型和标准化器"""
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    print(f"将使用的设备: {device}")
    model = MLPRegression(input_dim, output_dim, h1, h2, h3)
    try:
        model.load_state_dict(torch.load(model_path, map_location=device))
        model.to(device)
        model.eval()
        print(f"模型已成功从 {model_path} 加载。")
    except Exception as e:
        print(f"加载模型 {model_path} 时出错: {e}")
        return None, None, None, None
    try:
        scaler_X = joblib.load(x_scaler_path)
        scaler_y = joblib.load(y_scaler_path)
        print(f"Scaler 已成功从 {x_scaler_path} 和 {y_scaler_path} 加载。")
    except Exception as e:
        print(f"加载 Scaler 时出错 (如果训练未使用标准化，请忽略): {e}")
        # 如果训练时没用标准化，下面预测时也不用 scaler
        scaler_X, scaler_y = None, None
        # return None, None, None, None # 如果严格要求必须有 Scaler
    return model, scaler_X, scaler_y, device

# --- 5. 进行预测的函数 (与之前相同) ---
def predict_displacements(model, scaler_X, scaler_y, device, pressure_value):
    """使用加载的模型和 Scaler 对给定的压力值进行预测"""
    if model is None:
        print("模型未正确加载，无法进行预测。")
        return None

    input_p = np.array([[pressure_value]], dtype=np.float32)

    # 标准化输入 (如果使用了 Scaler)
    if scaler_X:
        input_p_scaled = scaler_X.transform(input_p)
    else:
        input_p_scaled = input_p # 没有 Scaler，直接使用原始输入

    input_tensor = torch.tensor(input_p_scaled, dtype=torch.float32).to(device)

    with torch.no_grad():
        predicted_scaled_tensor = model(input_tensor)

    predicted_scaled = predicted_scaled_tensor.cpu().numpy()

    # 反标准化输出 (如果使用了 Scaler)
    if scaler_y:
        predicted_original_scale = scaler_y.inverse_transform(predicted_scaled)
    else:
        predicted_original_scale = predicted_scaled # 没有 Scaler，直接使用模型输出

    return predicted_original_scale.flatten() # 返回 189 个位移值

# --- 6. 新增：加载初始坐标的函数 ---
def load_initial_coordinates(file_path, expected_nodes):
    """从文本文件加载初始坐标 (X0, Y0, Z0)"""
    try:
        coords = np.loadtxt(file_path, dtype=np.float32)
        if coords.shape == (expected_nodes, 3):
            print(f"成功从 {file_path} 加载 {coords.shape[0]} 个节点的初始坐标。")
            return coords # 返回形状为 (63, 3) 的数组
        else:
            print(f"错误：从 {file_path} 加载的坐标形状 {coords.shape} 与预期 ({expected_nodes}, 3) 不符。")
            return None
    except FileNotFoundError:
        print(f"错误：找不到初始坐标文件 {file_path}")
        return None
    except Exception as e:
        print(f"加载初始坐标文件 {file_path} 时出错: {e}")
        return None

# --- 7. 新增：计算变形后坐标的函数 ---
def calculate_deformed_coordinates(initial_coords, displacements_u):
    """计算变形后的坐标 X_def, Y_def, Z_def"""
    if initial_coords is None or displacements_u is None:
        return None
    if initial_coords.shape[0] * 3 != displacements_u.shape[0]:
        print(f"错误：初始坐标数量 ({initial_coords.shape[0]}) 与位移值数量 ({displacements_u.shape[0]}) 不匹配。")
        return None

    # 将 1D 的位移数组 (189,) 变形为 (63, 3)
    displacements_reshaped = displacements_u.reshape(initial_coords.shape[0], 3)

    # 计算变形后坐标
    deformed_coords = initial_coords + displacements_reshaped
    return deformed_coords # 返回形状为 (63, 3) 的数组

# --- 8. 绘制变形后曲线的函数 (强调点顺序的重要性) ---
# source: 1, 19, 21, 22, 23, 24
def plot_deformed_finger(deformed_coords, pressure_value):
    """
    使用 matplotlib 绘制 3D 变形曲线。
    通过查找最近邻点来确定连接顺序，而不是依赖于输入数组的顺序。
    """
    if deformed_coords is None:
        print("没有变形坐标可供绘制。")
        return
    num_points = deformed_coords.shape[0]
    if num_points < 2:
        print("至少需要两个点才能绘制连线。")
        return

    fig = plt.figure(figsize=(10, 8))
    ax = fig.add_subplot(111, projection='3d')

    # --- 最近邻连接逻辑 ---
    points_to_plot = deformed_coords.copy() # 复制一份，避免修改原始数据
    plotted_indices = set() # 记录已绘制或已使用的点的索引
    path_segments = [] # 存储要绘制的线段 [(start_coord, end_coord), ...]

    # 启发式选择起点：通常边缘点 X 坐标较小或较大，这里简单选第一个
    # (更好的方法可能需要基于几何特征选择)
    current_index = 0
    plotted_indices.add(current_index)
    remaining_indices = set(range(num_points)) - plotted_indices

    # 循环查找最近的未绘制点并连接
    while remaining_indices:
        current_coord = points_to_plot[current_index]
        min_dist_sq = float('inf')
        nearest_neighbor_index = -1

        # 在剩余点中查找最近邻
        for idx in remaining_indices:
            dist_sq = np.sum((current_coord - points_to_plot[idx])**2)
            if dist_sq < min_dist_sq:
                min_dist_sq = dist_sq
                nearest_neighbor_index = idx

        if nearest_neighbor_index != -1:
            # 找到了最近邻
            nearest_coord = points_to_plot[nearest_neighbor_index]
            # 添加线段到列表
            path_segments.append((current_coord, nearest_coord))

            # 更新当前点，标记为已用，从未用集合中移除
            current_index = nearest_neighbor_index
            plotted_indices.add(current_index)
            remaining_indices.remove(current_index)
        else:
            # 找不到更多可连接的点（例如，如果点集不连通）
            break # 退出循环
    # ------------------------

    # 绘制所有找到的线段
    if path_segments:
        for i, (start, end) in enumerate(path_segments):
            label = f'Deformed Edge (P={pressure_value})' if i == 0 else None # 只给第一段加标签
            ax.plot([start[0], end[0]], [start[1], end[1]], [start[2], end[2]],
                    linestyle='-', color='blue', label=label, zorder=1)
    else:
         # 如果没有线段（例如只有1个点），只画点
         ax.scatter(points_to_plot[:, 0], points_to_plot[:, 1], points_to_plot[:, 2],
                   color='red', s=20, label='Nodes', alpha=0.8, zorder=2)

    # 绘制所有节点本身
    ax.scatter(points_to_plot[:, 0], points_to_plot[:, 1], points_to_plot[:, 2],
               color='red', s=20, label='Nodes' if not path_segments else None, alpha=0.8, zorder=2)


    # 设置坐标轴标签和标题
    ax.set_xlabel('Global X')
    ax.set_ylabel('Global Y')
    ax.set_zlabel('Global Z')
    ax.set_title(f'Predicted Finger Deformation at P = {pressure_value}\n(Lines connect nearest neighbors)')

    # 尝试自动调整坐标轴范围
    try:
        x_def, y_def, z_def = points_to_plot[:, 0], points_to_plot[:, 1], points_to_plot[:, 2]
        max_range = np.array([x_def.max()-x_def.min(), y_def.max()-y_def.min(), z_def.max()-z_def.min()]).max()
        if max_range == 0: max_range = 1.0 # 防止范围为0
        max_range /= 1.8 # 调整缩放因子

        mid_x = (x_def.max()+x_def.min()) * 0.5
        mid_y = (y_def.max()+y_def.min()) * 0.5
        mid_z = (z_def.max()+z_def.min()) * 0.5
        ax.set_xlim(mid_x - max_range, mid_x + max_range)
        ax.set_ylim(mid_y - max_range, mid_y + max_range)
        ax.set_zlim(mid_z - max_range, mid_z + max_range)
    except Exception as e_axis:
        print(f"注意：自动调整坐标轴范围时出错: {e_axis}")

    ax.legend()
    plt.grid(True)
    plt.show() # 显示图形

# --- 9. 主程序 (修改后) ---
if __name__ == '__main__':
    # 加载模型和 Scaler
    model, scaler_X, scaler_y, device = load_prediction_components(
        MODEL_PATH, X_SCALER_PATH, Y_SCALER_PATH,
        INPUT_DIM, OUTPUT_DIM, HIDDEN_LAYER_1, HIDDEN_LAYER_2, HIDDEN_LAYER_3
    )

    # 预先加载初始坐标
    initial_coords = load_initial_coordinates(INITIAL_COORDS_PATH, NODE_COUNT)

    if model is not None and initial_coords is not None:
        while True:
            try:
                input_pressure_str = input("请输入要预测和可视化的气压 P 值 (输入 'quit' 退出): ")
                if input_pressure_str.lower() == 'quit':
                    break
                input_pressure = float(input_pressure_str)

                # 1. 进行预测，得到位移 U (189个值)
                predicted_u = predict_displacements(model, scaler_X, scaler_y, device, input_pressure)

                if predicted_u is not None:
                    print(f"\n成功预测压力 P = {input_pressure} 时的位移。")

                    # 将该变量写入Excel表格中

                    # 检查 predicted_u 是否有效
                    if predicted_u is not None and predicted_u.ndim == 1 and predicted_u.shape[0] == NODE_COUNT * 3:
                        try:
                            # 创建一个新的工作簿对象 (每次循环都会新建或准备覆盖)
                            workbook = Workbook()
                            # 获取活动工作表 (修正笔误：active 是属性，不是方法)
                            sheet = workbook.active
                            # 设置工作表标题 (修正笔误：tittle -> title)
                            sheet.title = "Displacement Comparison"

                            # 写入表头
                            sheet['A1'] = "Pressure"
                            sheet['B1'] = "NodeLabel"
                            sheet['C1'] = "X0"
                            sheet['D1'] = "Y0"
                            sheet['E1'] = "Z0"
                            sheet['F1'] = "Predicted_U1" # 修正表头更清晰
                            sheet['G1'] = "Predicted_U2" # 修正表头更清晰
                            sheet['H1'] = "Predicted_U3" # 修正表头更清晰

                            # 循环写入每个节点的数据
                            num_nodes = NODE_COUNT # 从脚本前面定义获取 [cite: 2]
                            for i in range(num_nodes):
                                # Excel 行从 1 开始，数据从第 2 行开始写
                                excel_row = i + 2

                                # 提取 U1, U2, U3
                                u1_value = predicted_u[i * 3]
                                u2_value = predicted_u[i * 3 + 1]
                                u3_value = predicted_u[i * 3 + 2]

                                # 写入 F, G, H 列 (列索引 F=6, G=7, H=8)
                                sheet.cell(row=excel_row, column=6, value=u1_value)
                                sheet.cell(row=excel_row, column=7, value=u2_value)
                                sheet.cell(row=excel_row, column=8, value=u3_value)

                                # (可选) 写入辅助信息到 A-E 列
                                sheet.cell(row=excel_row, column=1, value=input_pressure)       # Pressure 到 A 列
                                sheet.cell(row=excel_row, column=2, value=f"Node {i+1}")        # Node Label 到 B 列
                                if initial_coords is not None and i < initial_coords.shape[0]: # 检查初始坐标是否存在
                                    sheet.cell(row=excel_row, column=3, value=initial_coords[i, 0]) # X0 到 C 列
                                    sheet.cell(row=excel_row, column=4, value=initial_coords[i, 1]) # Y0 到 D 列
                                    sheet.cell(row=excel_row, column=5, value=initial_coords[i, 2]) # Z0 到 E 列

                            # 定义 Excel 文件名
                            excel_file_path = 'comparison.xlsx'
                            # 保存工作簿到当前脚本所在文件夹
                            workbook.save(excel_file_path)
                            print(f"预测结果已成功写入（或覆盖）到文件: {excel_file_path}")

                        except PermissionError:
                             print(f"错误：无法保存 Excel 文件 '{excel_file_path}'。请检查文件是否被其他程序打开或是否有写入权限。")
                        except Exception as e_excel:
                             print(f"写入 Excel 文件时发生错误: {e_excel}")
                    elif predicted_u is None:
                        print("预测结果 (predicted_u) 为空，无法写入 Excel。")
                    else:
                         print(f"错误：预测的位移数据形状 ({predicted_u.shape}) 不符合预期 ({NODE_COUNT * 3},)，无法写入 Excel。")

                    # 2. 计算变形后的坐标 (X, Y, Z) (这部分代码保持不变)
                    deformed_coords = calculate_deformed_coordinates(initial_coords, predicted_u)

                    # 2. 计算变形后的坐标 (X, Y, Z)
                    deformed_coords = calculate_deformed_coordinates(initial_coords, predicted_u)

                    # 3. 绘制变形后的曲线
                    if deformed_coords is not None:
                        print("正在生成 3D 可视化图形...")
                        plot_deformed_finger(deformed_coords, input_pressure)
                    else:
                        print("无法计算变形坐标，无法绘图。")
                else:
                    print("预测失败，无法绘图。")
                print("-" * 30)

            except ValueError:
                print("无效输入，请输入一个数字或 'quit'。")
            except Exception as e:
                print(f"处理过程中发生错误: {e}")

        print("程序退出。")
    else:
        if model is None:
             print("无法加载模型，程序退出。")
        if initial_coords is None:
             print(f"无法加载初始坐标，请确保 {INITIAL_COORDS_PATH} 文件存在且格式正确。程序退出。")