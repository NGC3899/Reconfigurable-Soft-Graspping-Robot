# -*- coding: utf-8 -*-

# =============================================================================
# Abaqus Python Script for Parametric Pressure Study
# Version: 6.4 (Fix Disp Data Check based on User V3.4) # <<< 版本号更新
# Extracts Max Contact Stresses (Sheet1 from stress_node_set_name)
# and Nodal Displacements (Sheet2 from disp_node_set_name) using NumPy check.
# Uses Node Sets for region specification. Strict syntax formatting applied.
# =============================================================================
#
# Description: Performs pressure iterations and extracts specified results.
# Author: [Your Name/Organization] - Modified based on user request
# Date: [Date]
# Prerequisites: (Updated)
#

# =============================================================================
# Import necessary modules
# =============================================================================
import os
import sys
import datetime
import traceback
# import math # 不需要 math

# --- 初始化 Debug Excel 日志 ---
target_directory = r'C:\Users\admin\Desktop\FEM'
debug_excel_file_name = 'Debug_Log.xlsx'
debug_excel_file = os.path.join(target_directory, debug_excel_file_name)
initial_log_success = False
wb_debug = None
ws_debug = None
log_init_dir_msg = None # 初始化

# --- 确保目录存在 (V6.4 严格结构) ---
print "DEBUG: Checking target directory: {}".format(target_directory) # 独立行
# if 块开始
if not os.path.exists(target_directory):
    # try 块开始
    try:
        print "DEBUG: Target directory does not exist. Attempting creation..." # 独立行
        os.makedirs(target_directory) # 独立行
        log_init_dir_msg = "INFO: Created target directory: {}".format(target_directory) # 独立行
        print log_init_dir_msg # 独立行
    # except OSError 块开始
    except OSError as e:
        print "FATAL ERROR: Cannot create target directory for logging: {}. Error: {}".format(target_directory, e) # 独立行
        sys.exit("Script aborted: cannot create log directory.") # 独立行
    # except Exception 块开始
    except Exception as e_create:
         print "FATAL ERROR: Unexpected error creating directory: {}. Error: {}".format(target_directory, e_create) # 独立行
         sys.exit("Script aborted: unexpected directory creation error.") # 独立行
# else 块开始
else:
    log_init_dir_msg = "DEBUG: Target directory already exists: {}".format(target_directory) # 独立行
    print log_init_dir_msg # 独立行

# --- 尝试导入 openpyxl 并初始化 Debug Excel (V6.4 严格结构) ---
try:
    print "DEBUG: Importing openpyxl..." # 独立行
    import openpyxl # 独立行
    from openpyxl import Workbook # 独立行
    print "DEBUG: Initializing Debug Workbook..." # 独立行
    wb_debug = Workbook() # 独立行
    ws_debug = wb_debug.active # 独立行
    ws_debug.title = 'DebugLog' # 独立行
    print "DEBUG: Appending headers to Debug Log..." # 独立行
    ws_debug.append(['Timestamp', 'Level', 'Message', 'Traceback']) # 独立行
    initial_log_success = True # 独立行
    print "DEBUG: Debug Excel log initialized." # 独立行
# except ImportError 块开始
except ImportError:
    print "FATAL ERROR: The 'openpyxl' library is required for logging but not found." # 独立行
    print "Please install it in your Abaqus Python environment." # 独立行
    print "Run: abaqus python -m pip install openpyxl" # 独立行
    sys.exit("Script aborted due to missing openpyxl for logging.") # 独立行
# except Exception 块开始
except Exception as e:
    print "FATAL ERROR: Failed to initialize Debug Excel Log ({}). Error: {}".format(debug_excel_file, e) # 独立行
    sys.exit("Script aborted: Cannot initialize debug log.") # 独立行

# --- 定义日志记录函数 ---
def log_to_excel(level, message, exc_obj=None):
    # if 块开始
    if not initial_log_success or ws_debug is None:
        print "LOG_FALLBACK: [{}] {}".format(level, message) # 独立行
        # if 块开始
        if exc_obj:
            print "LOG_FALLBACK_EXCEPTION: {}".format(exc_obj) # 独立行
        return # 独立行
    # try 块开始
    try:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        tb_str = "" # 独立行
        # if 块开始
        if exc_obj:
            # try 块开始
            try:
                tb_str = "".join(traceback.format_exception(type(exc_obj), exc_obj, exc_obj.__traceback__)) # 独立行
            # except 块开始
            except Exception as format_exc_err:
                try: # 尝试获取 repr
                    original_exc_repr = repr(exc_obj) # 独立行
                except Exception: # 获取 repr 也失败
                    original_exc_repr = "[Could not get repr() of original exception]" # 独立行
                # 记录包含类型和 repr 的回退信息
                tb_str = "Traceback formatting failed (Error: {}). Original Exception Type: <{}>, Original Exception Repr: {}".format(
                    format_exc_err, type(exc_obj).__name__, original_exc_repr
                ) # 多行赋值保持可读性
        ws_debug.append([str(timestamp), str(level), str(message), str(tb_str)]) # 独立行
    # except 块开始
    except Exception as log_e:
        print "ERROR writing to debug log: {}. Original message: [{}] {}".format(log_e, level, message) # 独立行

# --- 定义保存 Debug Excel 的函数 ---
def save_debug_log():
    # if 块开始
    if not initial_log_success or wb_debug is None:
         print "ERROR: Debug log was not initialized, cannot save." # 独立行
         return False # 独立行
    # try 块开始
    try:
        wb_debug.save(debug_excel_file) # 独立行
        print "Attempting to save debug log to {}".format(debug_excel_file) # 独立行
        return True # 独立行
    # except 块开始
    except Exception as save_e:
        print "ERROR: Failed to save debug log file '{}'. Error: {}".format(debug_excel_file, save_e) # 独立行
        return False # 独立行

# --- 开始记录日志 ---
if initial_log_success and log_init_dir_msg: # 独立行
    log_to_excel("INFO", log_init_dir_msg.replace("DEBUG: ", "Directory Info: ")) # 独立行
elif initial_log_success: # 独立行
    log_to_excel("INFO", "Debug Excel log active.") # 独立行

log_to_excel("DEBUG", "Importing Abaqus modules...")
from abaqus import *
from abaqusConstants import *
import odbAccess
import job
import load
import step
import assembly # 确保导入

# 尝试导入 numpy (严格分行)
try:
    import numpy as np
    log_to_excel("DEBUG", "Numpy imported successfully.")
except ImportError as e:
    err_msg = "The 'numpy' library is required..." # 省略
    log_to_excel("ERROR", err_msg, exc_obj=e) # 独立行
    save_debug_log() # 独立行
    sys.exit("Script aborted due to missing numpy.") # 独立行

log_to_excel("DEBUG", "Standard Python modules imported.")

# =============================================================================
# USER CONFIGURABLE PARAMETERS - MODIFY THESE VALUES
# =============================================================================
log_to_excel("DEBUG", "Setting user configurable parameters...")
cae_file_path = r'D:\softgripper2.cae' # <<< 确认路径正确
model_name = 'Model-1'
pressure_load_names = [
    'pressure1', 'pressure2', 'pressure3', 'pressure4a', 'pressure4b', 'pressure4c',
    'pressure4d', 'pressure4e', 'pressure4f', 'pressure5a', 'pressure5b', 'pressure5c',
    'pressure5d', 'pressure5e'
]
num_iterations = 2 # <<< 确认压力迭代次数
initial_pressure = 10000.0
pressure_increment = 1000.0
excel_file_name = 'Stress_And_Displacement_Results.xlsx' # <<< 修改了文件名
output_excel_file = os.path.join(target_directory, excel_file_name)

# --- ODB Result Extraction Parameters ---
# <<< 修改: 分别指定应力节点集和位移节点集 >>>
# !!! --- 请确保名称与你在 Abaqus/CAE 中定义的节点集名称完全一致 --- !!!
stress_node_set_name = 'NS_CONTACT_MAIN'  # <<< 用于提取应力最大值的节点集 (刚体)
disp_node_set_name = 'EDGE_FOR_DISP'   # <<< 用于提取位移的节点集 (软体)
# !!! ----------------------------------- !!!

step_name = 'Bending'
stress_output_variables = ["CPRESS", "CSHEAR1", "CSHEAR2"] # 应力变量列表
displacement_variable_name = 'U' # 位移变量名

log_to_excel("DEBUG", "User parameters set.")
log_to_excel("DEBUG", "Target output directory: {}".format(target_directory))
log_to_excel("DEBUG", "Output Excel file path: {}".format(output_excel_file))
log_to_excel("DEBUG", "Debug Excel file path: {}".format(debug_excel_file))
log_to_excel("DEBUG", "Pressure Iterations: {}".format(num_iterations))
log_to_excel("DEBUG", "Node Set for Stress: {}".format(stress_node_set_name))
log_to_excel("DEBUG", "Node Set for Displacement: {}".format(disp_node_set_name))
log_to_excel("DEBUG", "Stress variables: {}".format(stress_output_variables))
log_to_excel("DEBUG", "Displacement variable: {}".format(displacement_variable_name))


# =============================================================================
# Script Initialization
# =============================================================================
log_to_excel("INFO", "Starting Abaqus parametric pressure study script...")

# --- Initialize RESULT Excel Workbook with TWO Sheets ---
log_to_excel("DEBUG", "Initializing Result Excel workbook...")
stress_excel_headers = ["Iteration", "Pressure"] + ["Max " + var for var in stress_output_variables]
disp_excel_headers = ['Iteration', 'Pressure', 'NodeLabel', 'X0', 'Y0', 'Z0', 'U1', 'U2', 'U3'] # <<< 位移表头
wb_result = None
ws_stress = None # <<< 应力 Sheet
ws_disp = None   # <<< 位移 Sheet
try:
    wb_result = Workbook() # 独立行
    # 创建第一个 Sheet (应力)
    ws_stress = wb_result.active # 独立行
    ws_stress.title = 'Max Contact Stresses' # 独立行
    log_to_excel("DEBUG", "Appending headers to Stress sheet...") # 独立行
    ws_stress.append(stress_excel_headers) # 独立行
    log_to_excel("DEBUG", "Stress Excel headers written: {}".format(stress_excel_headers)) # 独立行
    # <<< 新增：创建第二个 Sheet (位移) >>>
    ws_disp = wb_result.create_sheet(title='Nodal Displacements') # 独立行
    log_to_excel("DEBUG", "Appending headers to Displacement sheet...") # 独立行
    ws_disp.append(disp_excel_headers) # <<< 写入位移表头
    log_to_excel("DEBUG", "Displacement Excel headers written: {}".format(disp_excel_headers)) # 独立行
except Exception as e:
    log_to_excel("ERROR", "Failed to initialize Result Excel workbook or sheets.", exc_obj=e) # 独立行
    save_debug_log() # 独立行
    sys.exit("Script aborted: Failed to initialize result excel.") # 独立行


# --- Open the CAE file and access the model ---
log_to_excel("DEBUG", "Attempting to open CAE file: {}".format(cae_file_path))
myModel = None
myAssembly = None # <<< 新增: 获取 Assembly 对象
initial_coords_dict = {} # <<< 新增: 存储初始坐标
try:
    openMdb(pathName=cae_file_path) # 独立行
    log_to_excel("INFO", "Successfully opened CAE file: {}".format(cae_file_path)) # 独立行
    myModel = mdb.models[model_name] # 独立行
    log_to_excel("INFO", "Accessed model: {}".format(model_name)) # 独立行
    myAssembly = myModel.rootAssembly # <<< 获取 Assembly 对象
    log_to_excel("INFO", "Accessed root assembly.") # 独立行

    # --- <<< 新增：获取位移节点集内节点的初始坐标 >>> ---
    log_to_excel("DEBUG", "Attempting to get initial coordinates for nodes in Node Set '{}' from MDB...".format(disp_node_set_name)) # <<< 使用 disp_node_set_name
    if myAssembly: # 检查 Assembly 对象是否存在
        try: # try 块开始
            # 假设节点集在 Assembly 层面定义
            node_set_mdb = myAssembly.sets[disp_node_set_name] # <<< 使用 disp_node_set_name
            log_to_excel("DEBUG", "Accessed Node Set '{}' from MDB.".format(disp_node_set_name)) # 独立行
            # 获取节点对象列表
            nodes_in_set_mdb = node_set_mdb.nodes # 独立行
            log_to_excel("DEBUG", "Found {} nodes in Node Set '{}' (MDB).".format(len(nodes_in_set_mdb), disp_node_set_name)) # 独立行
            # 遍历节点并存储坐标
            node_count = 0 # 计数器
            for node in nodes_in_set_mdb: # for 循环开始
                try: # try 块开始 (获取单个节点坐标)
                    initial_coords_dict[node.label] = node.coordinates # 独立行
                    node_count += 1 # 独立行
                except Exception as e_coord: # except 块开始 (获取单个节点坐标失败)
                    log_to_excel("WARNING", "Could not get coordinates for node with label {} in MDB.".format(node.label), exc_obj=e_coord) # 独立行
            log_to_excel("INFO", "Stored initial coordinates for {} nodes from set '{}'.".format(node_count, disp_node_set_name)) # 独立行
        except KeyError as e_set: # except 块开始 (MDB中找不到节点集)
            log_to_excel("ERROR", "Node Set '{}' for displacement not found in the Assembly of the CAE file '{}'. Please check the name.".format(disp_node_set_name, cae_file_path), exc_obj=e_set) # 独立行
            save_debug_log() # 独立行
            sys.exit("Script aborted: Node Set for initial coordinates not found in MDB.") # 独立行
        except Exception as e_get_init: # except 块开始 (其他获取初始坐标错误)
             log_to_excel("ERROR", "An error occurred while getting initial node coordinates from MDB.", exc_obj=e_get_init) # 独立行
             save_debug_log() # 独立行
             sys.exit("Script aborted: Failed to get initial coordinates.") # 独立行
    else: # else 块开始 (Assembly 对象获取失败)
         log_to_excel("ERROR", "Could not access Assembly object from model '{}'.".format(model_name)) # 独立行
         save_debug_log() # 独立行
         sys.exit("Script aborted: Assembly object not found.") # 独立行
    # --- <<< 新增结束 >>> ---

except MdbError as e: # except 块开始
    log_to_excel("ERROR", "Failed to open CAE file '{}'. MdbError.".format(cae_file_path), exc_obj=e) # 独立行
    save_debug_log() # 独立行
    sys.exit("Script aborted.") # 独立行
except KeyError as e: # except 块开始 (处理 Model 或 NodeSet 名称错误)
    error_message = str(e) # 独立行
    key_error_source = "Unknown" # 独立行
    if model_name in error_message: # 独立行
        key_error_source = "Model '{}'".format(model_name) # 独立行
    # <<< 修改: 同时检查两个 NodeSet 名称 >>>
    elif (disp_node_set_name in error_message or stress_node_set_name in error_message) and 'sets' in error_message:
        key_error_source = "Node Set '{}' or '{}'".format(disp_node_set_name, stress_node_set_name) # 独立行
    log_to_excel("ERROR", "KeyError accessing MDB object: {}. Details: {}".format(key_error_source, e), exc_obj=e) # 独立行
    save_debug_log() # 独立行
    sys.exit("Script aborted.") # 独立行
except Exception as e: # except 块开始
    log_to_excel("ERROR", "An unexpected error occurred opening CAE file or accessing initial objects.", exc_obj=e) # 独立行
    save_debug_log() # 独立行
    sys.exit("Script aborted.") # 独立行


# --- 更改当前工作目录到目标文件夹 ---
log_to_excel("DEBUG", "Changing current working directory to: {}".format(target_directory))
try:
    os.chdir(target_directory) # 独立行
    log_to_excel("INFO", "Current working directory changed successfully.") # 独立行
except OSError as e:
    log_to_excel("ERROR", "Could not change working directory to '{}'...".format(target_directory), exc_obj=e) # 独立行
    save_debug_log() # 独立行
    sys.exit("Script aborted due to chdir error.") # 独立行

# =============================================================================
# Main Simulation Loop (Single Loop for Pressure)
# =============================================================================
log_to_excel("INFO", "Entering main simulation loop (Pressure only)...") # <<< 修改日志

for i in range(num_iterations): # <<< 修改：使用 num_iterations
    iteration_number = i + 1 # <<< 修改：使用 iteration_number
    log_to_excel("INFO", "------------------ Starting Iteration {} of {} ------------------".format(iteration_number, num_iterations)) # <<< 修改

    # --- 计算压力 ---
    log_to_excel("DEBUG", "Calculating current pressure...") # 独立行
    current_pressure = initial_pressure + i * pressure_increment # 独立行
    log_to_excel("INFO", "Current Pressure: {:.2f} Pa".format(current_pressure)) # 独立行

    # --- 修改载荷 ---
    log_to_excel("DEBUG", "Updating pressure loads...") # 独立行
    try:
        if myModel is None:
            raise ValueError("MDB Model unavailable.") # 独立行
        for load_name in pressure_load_names:
            load_object = myModel.loads[load_name] # 独立行
            load_object.setValues(magnitude=current_pressure) # 独立行
        log_to_excel("DEBUG", "Pressure loads updated.") # 独立行
    except KeyError as e:
        log_to_excel("ERROR", "Pressure load name '{}' not found...".format(e), exc_obj=e) # 独立行
        save_debug_log() # 独立行
        sys.exit("Script aborted.") # 独立行
    except Exception as e:
        log_to_excel("ERROR", "Failed to update pressure loads.", exc_obj=e) # 独立行
        save_debug_log() # 独立行
        sys.exit("Script aborted.") # 独立行

    # --- 创建/运行 Job ---
    job_name = 'PressureStudy_Iter_{}'.format(iteration_number) # <<< 恢复简单 Job 名称
    log_to_excel("DEBUG", "Creating Abaqus job object: {}".format(job_name)) # 独立行
    myJob = None # 独立行
    try:
        if myModel is None:
            raise ValueError("MDB Model unavailable...") # 独立行
        myJob = mdb.Job(name=job_name, model=model_name,
                        description='Pressure study iteration {}'.format(iteration_number), # <<< 恢复简单描述
                        type=ANALYSIS, # ... 其他 Job 参数 ...
                        resultsFormat=ODB, multiprocessingMode=DEFAULT, numCpus=1, numDomains=1, numGPUs=0) # 独立行
        log_to_excel("DEBUG", "Job object created.") # 独立行
        log_to_excel("INFO", "Submitting job: {}".format(job_name)) # 独立行
        myJob.submit(consistencyChecking=OFF) # 独立行
        log_to_excel("DEBUG", "Job submission issued.") # 独立行
        log_to_excel("INFO", "Waiting for job: {}...".format(job_name)) # 独立行
        myJob.waitForCompletion() # 独立行
        log_to_excel("INFO", "Job {} completed.".format(job_name)) # 独立行
    except AbaqusException as e:
        log_to_excel("ERROR", "Abaqus job failed for {}.".format(job_name), exc_obj=e) # 独立行
        log_to_excel("WARNING", "Skipping results extraction.") # 独立行
        continue # 独立行
    except Exception as e:
        log_to_excel("ERROR", "Unexpected job error for {}.".format(job_name), exc_obj=e) # 独立行
        log_to_excel("WARNING", "Skipping results extraction.") # 独立行
        continue # 独立行

    # --- Post-processing: Extract results from ODB ---
    odb_path = job_name + '.odb' # 独立行
    log_to_excel("INFO", "Starting ODB post-processing for: {}".format(odb_path)) # 独立行
    odb = None # 独立行
    max_stress_values = {var: 0.0 for var in stress_output_variables} # <<< 应力结果字典
    log_to_excel("DEBUG", "Initialized max_stress_values: {}".format(max_stress_values)) # 独立行

    try: # ODB Main Try
        log_to_excel("DEBUG", "Attempting to open ODB file: {}".format(odb_path)) # 独立行
        odb_exists = os.path.exists(odb_path) # 独立行
        if not odb_exists: # if 块开始
             err_msg = "ODB file not found: {}. Job might have failed.".format(odb_path) # 调整消息 # 独立行
             log_to_excel("ERROR", err_msg) # 独立行
             # 记录错误并跳过此迭代的 ODB 处理
             node_set_region_stress = None # 确保后面跳过
             node_set_region_disp = None   # 确保后面跳过
             last_frame = None             # 确保后面跳过
             log_to_excel("WARNING", "Skipping ODB processing for this iteration.") # 独立行
        else: # else 块开始 (ODB 文件存在)
            odb = odbAccess.openOdb(path=odb_path, readOnly=True) # 独立行
            log_to_excel("INFO", "Successfully opened ODB file: {}".format(odb_path)) # 独立行

            # --- <<< 修改：分别获取应力节点集和位移节点集对象 >>> ---
            node_set_region_stress = None # 应力区域对象
            node_set_region_disp = None   # 位移区域对象
            odb_node_sets = odb.rootAssembly.nodeSets # 获取一次即可

            # 获取应力节点集
            log_to_excel("DEBUG", "Attempting to access Stress NodeSet: '{}' in ODB...".format(stress_node_set_name)) # 独立行
            try: # Stress NodeSet Try
                if stress_node_set_name in odb_node_sets.keys(): # if 块开始
                    node_set_region_stress = odb_node_sets[stress_node_set_name] # 独立行
                    log_to_excel("INFO", "Accessed Stress NodeSet: {}".format(stress_node_set_name)) # 独立行
                else: # else 块开始
                    log_to_excel("ERROR", "Could not find Stress NodeSet '{}' in ODB.".format(stress_node_set_name)) # 独立行
            except Exception as e_stress_ns: # Stress NodeSet Except
                log_to_excel("ERROR", "Error accessing Stress NodeSet '{}'.".format(stress_node_set_name), exc_obj=e_stress_ns) # 独立行

            # 获取位移节点集
            log_to_excel("DEBUG", "Attempting to access Displacement NodeSet: '{}' in ODB...".format(disp_node_set_name)) # 独立行
            try: # Disp NodeSet Try
                if disp_node_set_name in odb_node_sets.keys(): # if 块开始
                    node_set_region_disp = odb_node_sets[disp_node_set_name] # 独立行
                    log_to_excel("INFO", "Accessed Displacement NodeSet: {}".format(disp_node_set_name)) # 独立行
                else: # else 块开始
                    log_to_excel("ERROR", "Could not find Displacement NodeSet '{}' in ODB.".format(disp_node_set_name)) # 独立行
            except Exception as e_disp_ns: # Disp NodeSet Except
                log_to_excel("ERROR", "Error accessing Displacement NodeSet '{}'.".format(disp_node_set_name), exc_obj=e_disp_ns) # 独立行
            # --- <<< 修改结束 >>> ---


            # --- Access the last frame ---
            analysis_step = None # 独立行
            last_frame = None # 独立行
            # if 块开始 (只要应力或位移节点集之一有效就可以尝试获取帧)
            if node_set_region_stress or node_set_region_disp:
                log_to_excel("DEBUG", "Attempting to access step: {}".format(step_name)) # 独立行
                try: # Step/Frame Try
                    analysis_step = odb.steps[step_name] # 独立行
                    log_to_excel("DEBUG", "Step '{}' accessed.".format(step_name)) # 独立行
                    if analysis_step.frames: # if 块开始
                        last_frame = analysis_step.frames[-1] # 独立行
                        log_to_excel("DEBUG", "Accessed last frame (ID: {})".format(last_frame.frameId)) # 独立行
                    else: # else 块开始
                        log_to_excel("WARNING", "Step '{}' has no frames.".format(step_name)) # 独立行
                        last_frame = None # 独立行
                except KeyError as e: # except 块开始
                    log_to_excel("ERROR", "Step '{}' not found.".format(step_name), exc_obj=e) # 独立行
                    last_frame = None # 独立行
                except IndexError as e: # except 块开始
                    log_to_excel("ERROR", "Error accessing frames.", exc_obj=e) # 独立行
                    last_frame = None # 独立行
            else: # else 块开始
                log_to_excel("WARNING", "Skipping step/frame access: Both NodeSets invalid.") # 独立行

            # --- <<< 修改：分离应力提取和位移提取 >>> ---
            if last_frame: # 确保帧有效

                # --- 1. 提取最大应力值 ---
                if node_set_region_stress: # 仅当应力节点集有效时执行
                    log_to_excel("DEBUG", "Extracting max stress values from NodeSet '{}'...".format(stress_node_set_name)) # 独立行
                    for variable_name in stress_output_variables: # for 循环开始 (应力变量)
                        log_to_excel("DEBUG", "--- Processing stress var: {} ---".format(variable_name)) # 独立行
                        max_val = 0.0 # 独立行
                        data_found = False # 独立行
                        subset = None # 独立行
                        try: # Variable processing Try
                             fieldOutput = last_frame.fieldOutputs[variable_name] # 独立行
                             subset = fieldOutput.getSubset(region=node_set_region_stress, position=NODAL) # <<< 使用 stress_node_set_region
                             if subset is None: # if 块开始
                                 log_to_excel("WARNING","Subset is None for stress var {}.".format(variable_name)) # 独立行
                                 max_val=0.0 # 独立行
                                 data_found=False # 独立行
                             else: # else 块开始
                                 # ... (bulk/values data extraction for stress - 不变, 确保分行) ...
                                 if hasattr(subset, 'bulkDataBlocks') and subset.bulkDataBlocks:
                                     try:
                                         block_data = subset.bulkDataBlocks[0].data
                                         if block_data is not None and len(block_data) > 0:
                                             data_array = np.array(block_data).flatten()
                                             if data_array.size > 0: max_val = np.max(data_array); data_found = True; log_to_excel("INFO", "Max {} (bulk): {:.4e}".format(variable_name, max_val))
                                             else: log_to_excel("WARNING", "bulkData empty '{}'.".format(variable_name))
                                         else: log_to_excel("WARNING", "bulkData[0].data empty '{}'.".format(variable_name))
                                     except Exception as bulk_err: log_to_excel("ERROR", "Error processing bulkData for {}.".format(variable_name), exc_obj=bulk_err)
                                 if not data_found and hasattr(subset, 'values') and subset.values:
                                     try:
                                         valid_data = [fv.data for fv in subset.values if hasattr(fv, 'data') and isinstance(fv.data, (int, float))]
                                         if valid_data: max_val = max(valid_data); data_found = True; log_to_excel("INFO", "Max {} (values): {:.4e}".format(variable_name, max_val))
                                         else: log_to_excel("WARNING", "No valid numeric data in values for {}.".format(variable_name))
                                     except Exception as val_err: log_to_excel("ERROR", "Error iterating values for {}.".format(variable_name), exc_obj=val_err)
                                 if not data_found: log_to_excel("WARNING", "Could not extract stress data for {}. Max=0.0.".format(variable_name)); max_val = 0.0
                        except KeyError as e: # except 块开始
                             log_to_excel("WARNING", "Field output '{}' not found.".format(variable_name)) # 独立行
                             max_val = 0.0 # 独立行
                        except Exception as e: # except 块开始
                             log_to_excel("ERROR", "Unexpected error processing stress {}. Max=0.0.".format(variable_name), exc_obj=e) # 独立行
                             max_val = 0.0 # 独立行

                        # Store result for this stress variable
                        log_to_excel("DEBUG", "Storing final max stress {:.4e} for '{}'".format(max_val, variable_name)) # 独立行
                        max_stress_values[variable_name] = max_val # <<< 存到 max_stress_values
                    # End stress variable loop
                    log_to_excel("DEBUG", "Finished extracting max stress values.") # 独立行
                else: # else for if stress_node_set_region
                    log_to_excel("WARNING", "Skipping stress extraction because stress NodeSet was not found/accessed.") # 独立行

                # --- 2. 提取节点位移值 ---
                if node_set_region_disp: # 仅当位移节点集有效时执行
                    log_to_excel("DEBUG", "Extracting nodal displacement values ('{}') from NodeSet '{}'...".format(displacement_variable_name, disp_node_set_name)) # 独立行
                    try: # try for displacement extraction
                        disp_field = last_frame.fieldOutputs[displacement_variable_name] # 获取 'U' 场输出
                        log_to_excel("DEBUG", "Field output '{}' obtained.".format(displacement_variable_name)) # 独立行
                        # 获取 'U' 在节点集上的子集
                        disp_subset = disp_field.getSubset(region=node_set_region_disp, position=NODAL) # <<< 使用 disp_node_set_region
                        log_to_excel("DEBUG", "Subset obtained for displacement.") # 独立行

                        if disp_subset and hasattr(disp_subset, 'values') and disp_subset.values: # 检查子集和 values 是否有效
                            log_to_excel("DEBUG", "Iterating through {} displacement values...".format(len(disp_subset.values))) # 独立行
                            # 遍历每个节点的值
                            for fv in disp_subset.values: # for 循环开始
                                try: # try for processing one FieldValue
                                    node_label = fv.nodeLabel # 获取节点标签
                                    # <<< V6.4 修改: 使用 NumPy 检查位移数据 >>>
                                    if hasattr(fv, 'data'): # 首先检查 data 属性是否存在
                                        try: # 尝试转换为 NumPy 数组并检查大小
                                            disp_array = np.array(fv.data) # 尝试转换
                                            if disp_array.size == 3: # 检查元素数量是否为 3
                                                u1 = disp_array.item(0) # 使用 .item() 获取标量值
                                                u2 = disp_array.item(1)
                                                u3 = disp_array.item(2)
                                                # 从 MDB 获取初始坐标
                                                initial_coords = initial_coords_dict.get(node_label, (None, None, None)) # 独立行
                                                x0 = initial_coords[0] # 独立行
                                                y0 = initial_coords[1] # 独立行
                                                z0 = initial_coords[2] # 独立行
                                                # 准备写入 Excel 的行数据
                                                disp_row = [iteration_number, current_pressure, node_label, x0, y0, z0, u1, u2, u3] # 独立行
                                                # 写入位移 Sheet
                                                if ws_disp: # 检查 ws_disp 是否有效
                                                    ws_disp.append(disp_row) # 独立行
                                                else: # 独立行
                                                    log_to_excel("ERROR", "Displacement sheet (ws_disp) is not available!") # 独立行
                                            else: # else 块开始 (Numpy 数组大小不对)
                                                log_to_excel("WARNING", "Displacement data for node {} has unexpected size after numpy conversion: {}. Data: {}".format(node_label, disp_array.size, fv.data)) # 独立行
                                        except Exception as np_conv_err: # except 块开始 (转换或处理 NumPy 数组失败)
                                             log_to_excel("WARNING", "Could not process displacement data for node {} as numpy array.".format(node_label), exc_obj=np_conv_err) # 独立行
                                    else: # else 块开始 (fv 没有 data 属性)
                                        log_to_excel("WARNING", "FieldValue object for node {} missing 'data' attribute.".format(node_label)) # 独立行
                                except AttributeError as e_fv_attr: # except 块开始 (访问 fv 属性失败)
                                    log_to_excel("WARNING", "Error accessing attributes (nodeLabel/data) for a FieldValue object.", exc_obj=e_fv_attr) # 独立行
                                # 不再需要之前的 IndexError/TypeError，因为包含在 Exception 里了
                            log_to_excel("DEBUG", "Finished iterating displacement values.") # 独立行
                        else: # else 块开始 (disp_subset 无效或无 values)
                            log_to_excel("WARNING", "Could not get valid displacement subset or values for NodeSet '{}'.".format(disp_node_set_name)) # 独立行

                    except KeyError: # except 块开始 (找不到 'U' 场输出)
                        log_to_excel("ERROR", "Field output '{}' not found in ODB frame. Cannot extract displacements.".format(displacement_variable_name)) # 独立行
                    except Exception as e_disp: # except 块开始 (提取位移的其他错误)
                        log_to_excel("ERROR", "An unexpected error occurred during displacement extraction.", exc_obj=e_disp) # 独立行
                    # --- 位移提取结束 ---
                else: # else for if disp_node_set_region
                    log_to_excel("WARNING", "Skipping displacement extraction because displacement NodeSet was not found/accessed.") # 独立行

            else: # else block for "if last_frame:"
                log_to_excel("WARNING", "Skipping all ODB data extraction for Iteration {}: Frame object is invalid.".format(iteration_number)) # 独立行

        # except blocks for outer ODB processing try (ODB 文件存在时的处理)
    except odbAccess.OdbError as e:
        log_to_excel("ERROR", "An OdbError occurred opening/processing ODB '{}'.".format(odb_path), exc_obj=e) # 独立行
        log_to_excel("WARNING", "Skipping results extraction...") # 独立行
    except Exception as e:
        log_to_excel("ERROR", "Unexpected error during ODB processing for '{}'.".format(odb_path), exc_obj=e) # 独立行
        log_to_excel("WARNING", "Skipping results extraction...") # 独立行
    # finally block for outer ODB processing try
    finally:
         if odb: # if 块开始
             log_to_excel("DEBUG", "Attempting to close ODB file: {}".format(odb_path)) # 独立行
             try: # 关闭 ODB 的 try
                 odb.close() # 独立行
                 log_to_excel("DEBUG", "Closed ODB file: {}".format(odb_path)) # 独立行
             except Exception as close_e: # 关闭 ODB 的 except
                 log_to_excel("ERROR", "Error closing ODB file '{}'.".format(odb_path), exc_obj=close_e) # 独立行
         else: # else 块开始
              log_to_excel("DEBUG", "ODB was not opened or error before closing.") # 独立行

    # --- <<< 修改：写入应力数据到应力 Sheet >>> ---
    log_to_excel("DEBUG", "Preparing row data for Stress Excel...") # 独立行
    stress_row_data = [iteration_number, current_pressure] # 独立行
    for var in stress_output_variables: # for 循环开始
        stress_row_data.append(max_stress_values.get(var, 0.0)) # 独立行
    log_to_excel("DEBUG", "Attempting to write row to Stress Excel: {}".format(stress_row_data)) # 独立行
    try: # try 块开始
        if ws_stress: # <<< 使用 ws_stress >>>
            ws_stress.append(stress_row_data) # 独立行
            log_to_excel("DEBUG", "Row successfully appended to Stress Excel worksheet.") # 独立行
        else: # else 块开始
            log_to_excel("ERROR","Stress worksheet (ws_stress) unavailable.") # 独立行
    except Exception as e: # except 块开始
        log_to_excel("ERROR", "Failed to append data to Stress Excel worksheet.", exc_obj=e) # 独立行

    log_to_excel("INFO", "------------------ Iteration {} completed ------------------".format(iteration_number)) # 独立行

# =============================================================================
# Script Finalization
# =============================================================================
log_to_excel("INFO", "Script execution finished main loop.") # 独立行
# --- 保存 Result Excel (包含两个 Sheet) ---
# ... (代码不变, 确保 except 分行) ...
log_to_excel("INFO", "Attempting to save final Result Excel workbook to: {}".format(output_excel_file))
try:
    if wb_result:
        wb_result.save(filename=output_excel_file); #... success logs ...
        log_to_excel("INFO", "=====================================================")
        log_to_excel("INFO", "Script finished successfully (Result Excel saved)!")
        log_to_excel("INFO", "All {} iterations appear completed.".format(num_iterations))
        log_to_excel("INFO", "Results potentially saved to: {}".format(os.path.abspath(output_excel_file)))
        log_to_excel("INFO", "=====================================================")
        log_to_excel("DEBUG", "Result Excel workbook saved successfully.")
    else: log_to_excel("ERROR","Result workbook unavailable.")
except Exception as e: log_to_excel("ERROR", "Failed to save Result Excel '{}'...".format(output_excel_file), exc_obj=e)
# --- 保存 Debug Excel ---
log_to_excel("INFO", "Attempting to save Debug Log Excel file...") # 独立行
saved_debug = save_debug_log() # 独立行
log_to_excel("INFO", "Script execution complete. Debug log save status: {}".format(saved_debug)) # 独立行
# --- 最后再次尝试保存 Debug Log ---
if not saved_debug: # 独立行
    log_to_excel("INFO", "Retrying to save Debug Log Excel file at the very end...") # 独立行
    save_debug_log() # 独立行
# 脚本结束